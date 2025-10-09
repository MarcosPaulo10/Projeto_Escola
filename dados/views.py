import io
from django.http import HttpResponse
from django.apps import apps
from django.db.models import ForeignKey, ManyToManyField
from django.db.models.fields.files import FieldFile 
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.shortcuts import render
from cadastros.models import Aluno, Responsavel, Responsavel_Aluno
from cadastros.models import Faturamento
from django.contrib.auth.decorators import login_required

@login_required
def export_all_tables_to_xlsx(request):
    output = io.BytesIO()
    workbook = Workbook()
    workbook.remove(workbook.active)

    excluded_apps = ['admin', 'auth', 'contenttypes', 'sessions', 'messages', 'staticfiles']
    all_models = apps.get_models()

    for model in all_models:
        app_label = model._meta.app_label
        model_name = model._meta.model_name

        if app_label in excluded_apps:
            continue

        # Criação da Planilha "Raw"
        sheet_raw_name = f"{model_name}_raw"[:31]
        sheet_raw = workbook.create_sheet(title=sheet_raw_name)

        fields = [field for field in model._meta.get_fields() if field.concrete]
        header_raw = [field.name for field in fields]
        sheet_raw.append(header_raw)

        for obj in model.objects.all():
            row = []
            for field in fields:
                value = getattr(obj, field.name)
                
                if isinstance(field, ForeignKey) and value is not None:
                    value = value.pk
                elif isinstance(field, ManyToManyField):
                    value = ", ".join(str(pk) for pk in value.values_list('pk', flat=True))
                elif isinstance(value, FieldFile):
                    value = str(value) if value else ""
                
                row.append(value)
            sheet_raw.append(row)


        # Criação da Planilha "Resolved"
        sheet_resolved_name = f"{model_name}_resolved"[:31]
        sheet_resolved = workbook.create_sheet(title=sheet_resolved_name)

        header_resolved = header_raw
        sheet_resolved.append(header_resolved)

        for obj in model.objects.all():
            row = []
            for field in fields:
                value = getattr(obj, field.name)

                if field.is_relation:
                    if value is not None:
                        if isinstance(field, ManyToManyField):
                            value = ", ".join(str(related_obj) for related_obj in value.all())
                        else:
                            value = str(value)
                elif isinstance(value, FieldFile):
                    value = str(value) if value else ""

                row.append(value)
            sheet_resolved.append(row)

    workbook.save(output)
    output.seek(0)

    filename = "database_export.xlsx"
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'

    return response

@login_required
def export_page_view(request):
    return render(request, 'export_page.html')

@login_required
def gerar_planilha_familias_view(request):
    # Cria um Workbook do Excel em memória
    wb = Workbook()
    ws = wb.active
    ws.title = "Agrupamento de Famílias"

    # Define o cabeçalho da planilha
    ws.append(['ID_Familia_Sugerido', 'Membros', 'IDs Responsáveis', 'IDs Alunos'])

    processados_ids = set()
    id_familia = 1

    for responsavel in Responsavel.objects.all():
        if responsavel.id in processados_ids:
            continue

        alunos_diretos_ids = set(Responsavel_Aluno.objects.filter(responsavel=responsavel).values_list('aluno_id', flat=True))
        if not alunos_diretos_ids:
            continue

        todos_responsaveis_ids = set(Responsavel_Aluno.objects.filter(aluno_id__in=alunos_diretos_ids).values_list('responsavel_id', flat=True))
        todos_alunos_ids = set(Responsavel_Aluno.objects.filter(responsavel_id__in=todos_responsaveis_ids).values_list('aluno_id', flat=True))

        responsaveis_familia = Responsavel.objects.filter(id__in=todos_responsaveis_ids)
        alunos_familia = Aluno.objects.filter(id__in=todos_alunos_ids)

        # Formata a string de membros
        membros_formatados = []
        for r in responsaveis_familia:
            membros_formatados.append(f"Rsp_{r.nome}")
        for a in alunos_familia:
            membros_formatados.append(f"Aln_{a.nome}")
        
        linha_membros = "|".join(membros_formatados)
        
        # Adiciona a linha na planilha
        ws.append([
            id_familia, 
            linha_membros,
            ", ".join(map(str, todos_responsaveis_ids)), # IDs para conferência
            ", ".join(map(str, todos_alunos_ids))      # IDs para conferência
        ])

        processados_ids.update(todos_responsaveis_ids)
        id_familia += 1
    
    # Prepara a resposta HTTP para download
    # Cria um buffer de bytes em memória para salvar o arquivo
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0) # Retorna o cursor para o início do buffer

    # Cria a resposta HTTP, definindo os headers para forçar o download
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="agrupamento_familias.xlsx"'

    return response

@login_required
def exportar_faturamento_view(request):
    # Obter os parâmetros de filtro da URL (via GET)
    mes_filtro = request.GET.get('mes')
    status_filtro = request.GET.get('status_pagamento')

    # Inicia a consulta base
    faturamentos = Faturamento.objects.select_related('familia__responsavel_financeiro').all()

    # Aplica os filtros ao queryset se eles foram fornecidos
    if mes_filtro and mes_filtro.isdigit() and int(mes_filtro) > 0:
        faturamentos = faturamentos.filter(mes_referencia__month=int(mes_filtro))

    if status_filtro == 'pago':
        faturamentos = faturamentos.filter(pago=True)
    elif status_filtro == 'nao_pago':
        faturamentos = faturamentos.filter(pago=False)

    output = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Relatorio de Faturamento"

    # Define o NOVO cabeçalho da planilha
    header = [
        'Responsável Financeiro',       # Coluna A
        'CPF/CNPJ do Responsável',      # Coluna B
        'Valor Pago',                   # Coluna C
        'Gerou NFe?',                   # NOVA COLUNA D
        'Mês de Referência',            # Coluna E
        'Status do Pagamento'           # Coluna F
    ]
    sheet.append(header)

    # Itera sobre os faturamentos JÁ FILTRADOS e adiciona as linhas
    for f in faturamentos:
        responsavel = f.familia.responsavel_financeiro
        
        nome_responsavel = responsavel.nome if responsavel else "Não definido"
        cpf_cnpj_responsavel = responsavel.cpf_cnpj if responsavel else "Não definido"

        valor_pago = f.valor_pago if f.valor_pago is not None else 0.00
        mes_referencia = f.mes_referencia.month
        status_pagamento = "Pago" if f.pago else "Em Aberto"

        # Cria a linha com a NOVA COLUNA
        row = [
            nome_responsavel,
            cpf_cnpj_responsavel,
            valor_pago,
            "Não",  # Valor fixo para a nova coluna
            mes_referencia,
            status_pagamento
        ]
        sheet.append(row)

    workbook.save(output)
    output.seek(0)

    filename = "relatorio_faturamento.xlsx"
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'

    return response
